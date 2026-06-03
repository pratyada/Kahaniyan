"""Read the 30-day plan spreadsheet."""
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from .config import CONTENT_DIR


@dataclass
class InstagramDay:
    day: int
    date: str
    title: str
    reel_concept: str
    voiceover: str
    caption: str
    cta: str
    hashtags: str
    status: str
    variety: bool = False


@dataclass
class XPostDay:
    day: int
    date: str
    pillar: str
    format: str
    post: str
    status: str


def clean(val):
    """Strip smart quotes and whitespace."""
    if not val:
        return ""
    s = str(val).strip()
    for old, new in [("\u201c", '"'), ("\u201d", '"'), ("\u2018", "'"), ("\u2019", "'")]:
        s = s.replace(old, new)
    return s


def load_plan(path: Path = None):
    path = path or CONTENT_DIR / "plan.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)

    # Instagram sheet
    ig_sheet = None
    for name in wb.sheetnames:
        if "instagram" in name.lower():
            ig_sheet = wb[name]
            break

    instagram = []
    if ig_sheet:
        for row in ig_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            title = clean(row[2])
            variety = "[VARIETY]" in title or "[variety]" in title.lower()
            instagram.append(InstagramDay(
                day=int(row[0]),
                date=clean(row[1]),
                title=title.replace("[VARIETY]", "").strip(),
                reel_concept=clean(row[3]),
                voiceover=clean(row[4]),
                caption=clean(row[5]),
                cta=clean(row[6]),
                hashtags=clean(row[7]),
                status=clean(row[8]) or "To do",
                variety=variety,
            ))

    # Platform X sheet
    x_sheet = None
    for name in wb.sheetnames:
        if "platform x" in name.lower() or "✖" in name:
            x_sheet = wb[name]
            break

    xposts = []
    if x_sheet:
        for row in x_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            xposts.append(XPostDay(
                day=int(row[0]),
                date=clean(row[1]),
                pillar=clean(row[2]),
                format=clean(row[3]),
                post=clean(row[4]),
                status=clean(row[5]) or "To do",
            ))

    wb.close()
    return instagram, xposts
